# =============================================================================
# LOADER.PY — SharePoint Write Functions
# Handles: save raw file, save formatted recon file, upload log
# =============================================================================

import requests
import pandas as pd
import urllib3
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

import config
from logger import log

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


# =============================================================================
# SAVE RAW CLASSIFIED FILE TO SHAREPOINT
# =============================================================================

def save_raw_file(df, token, sp_site_id, start_date, start_time, end_date, end_time):
    """Save classified emails dataframe to SharePoint raw_file folder"""

    file_name = (
        f"classified_emails_{start_date}_{start_time.replace(':','')}"
        f"_to_{end_date}_{end_time.replace(':','')}_IST.xlsx"
    )

    try:
        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Classified Emails", index=False)
        buffer.seek(0)

        url = (
            f"https://graph.microsoft.com/v1.0/sites/{sp_site_id}"
            f"/drives/{config.SP_DRIVE_ID}/root:{config.SP_RAW_FOLDER}/{file_name}:/content"
        )
        headers_upload = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        }
        resp = requests.put(url, headers=headers_upload, data=buffer.read(), verify=False)

        if resp.status_code in [200, 201]:
            log(f"Raw file saved : {config.SP_RAW_FOLDER}/{file_name}")
        else:
            log(f"Raw file save failed : {resp.status_code}", "ERROR")

    except Exception as e:
        log(f"save_raw_file error : {str(e)}", "ERROR")


# =============================================================================
# BUILD RECON DATAFRAME
# =============================================================================

def build_recon_df(df_live):
    """Build the recon dataframe from classified emails"""

    df_recon = pd.DataFrame()
    df_recon["From"]                 = df_live["sender_name"]
    df_recon["to_recipients"]        = df_live["to_recipients"]
    df_recon["Subject"]              = df_live["subject"]
    df_recon["Received Time"]        = df_live["time"]
    df_recon["Received Date"]        = pd.to_datetime(df_live["date"]).dt.strftime("%d-%b-%y")
    df_recon["Owner"]                = ""
    df_recon["Action"]               = ""
    df_recon["Status"]               = ""
    df_recon["actual_body"]          = df_live["actual_body"]
    df_recon["Comments"]             = df_live["predicted_class"]
    df_recon["Checked by"]           = ""
    df_recon["TAT status"]           = ""
    df_recon["Communication status"] = ""
    df_recon["case_number"]          = df_live["case_number"]

    # Sort by EST time
    df_recon["sort_dt"] = pd.to_datetime(
        df_live["date"].astype(str) + " " +
        df_live["time"].str.extract(r"(\d{2}:\d{2} [AP]M)")[0],
        format="%Y-%m-%d %I:%M %p",
        errors="coerce"
    )
    df_recon = df_recon.sort_values("sort_dt", ascending=True).reset_index(drop=True)
    df_recon = df_recon.drop(columns=["sort_dt"])

    log(f"Recon dataframe built : {len(df_recon)} rows")
    return df_recon


# =============================================================================
# SAVE FORMATTED RECON FILE TO SHAREPOINT
# =============================================================================

def save_recon_file(df_recon, token, sp_site_id, start_date, start_time, end_date, end_time):
    """Save formatted recon file with header styling to SharePoint landing zone"""

    file_name = (
        f"email_recon_{start_date}_{start_time.replace(':','')}"
        f"_to_{end_date}_{end_time.replace(':','')}_IST.xlsx"
    )

    try:
        # ── Write to buffer ───────────────────────────────────────────────────
        buffer = BytesIO()
        df_recon.to_excel(buffer, index=False)
        buffer.seek(0)

        # ── Apply formatting ──────────────────────────────────────────────────
        wb = load_workbook(buffer)
        ws = wb.active

        header_fill  = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
        header_font  = Font(bold=True, color="000000")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border  = Border(
            left=Side(style="thin"),  right=Side(style="thin"),
            top=Side(style="thin"),   bottom=Side(style="thin")
        )

        for cell in ws[1]:
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = header_align
            cell.border    = thin_border

        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 4, 40)

        ws.freeze_panes = "A2"

        # ── Save formatted to new buffer ──────────────────────────────────────
        formatted_buffer = BytesIO()
        wb.save(formatted_buffer)
        formatted_buffer.seek(0)

        # ── Upload to SharePoint ──────────────────────────────────────────────
        url = (
            f"https://graph.microsoft.com/v1.0/sites/{sp_site_id}"
            f"/drives/{config.SP_DRIVE_ID}/root:{config.SP_RECON_FOLDER}/{file_name}:/content"
        )
        headers_upload = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        }
        resp = requests.put(url, headers=headers_upload, data=formatted_buffer.read(), verify=False)

        if resp.status_code in [200, 201]:
            log(f"Recon file saved : {config.SP_RECON_FOLDER}/{file_name}")
        else:
            log(f"Recon file save failed : {resp.status_code}", "ERROR")

    except Exception as e:
        log(f"save_recon_file error : {str(e)}", "ERROR")
